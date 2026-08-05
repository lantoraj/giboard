# -*- coding: utf-8 -*-
"""
CZ-DRG extraction for GI-Board.

Two independent sources (both ÚZIS / MZ ČR open data, CC BY 4.0):

1) Národní ukazatele akutní lůžkové péče dle CZ-DRG  (data.mzcr.cz, CSV, rok 2022)
   Four hierarchy levels: MDC -> kategorie -> báze -> skupiny.
   Clinical + ECONOMIC indicators (costs, case-mix) – the only cost data in the
   whole dashboard.

2) Hospitalizační případy dle IČZ a CZ-DRG  (nzip.cz, XLSX, roky 2021-2024)
   Case counts per provider (IČZ) x DRG unit. Wide matrix: rows = DRG units,
   columns = providers -> pivoted to a sparse map here.

Note the two sources differ: (1) is national and single-year with costs,
(2) is per-provider over four years without costs. They share DRG codes.

Output: public/data/drg/
   indicators_<level>.json   national indicators per hierarchy level
   providers.json            sparse per-provider case counts (MDC + báze)
   meta.json                 years, sources, counts
"""
import csv, io, json, os, sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(r"C:\Users\lanto\OneDrive\Plocha\Vykony\gastro-dashboard")
SRC = Path(r"C:\Users\lanto\AppData\Local\Temp\claude\C--Users-lanto-OneDrive-Plocha-Vykony\76ed996e-92c6-4cff-b418-31e19375d18a\scratchpad")
OUT = ROOT / "public" / "data" / "drg"
OUT.mkdir(parents=True, exist_ok=True)

INDICATOR_YEAR = "2022"          # the v4 indicator files cover 2022
PROVIDER_YEARS = ["2021", "2022", "2023", "2024"]

# Keep provider detail at MDC + báze level; DRG skupina (1753 units) would
# roughly triple the payload for little added value in a provider ranking.
PROVIDER_LEVELS = {"MDC", "DRG báze"}

LEVELS = {                       # output name -> source csv
    "mdc":       "drg_mdc.csv",
    "kategorie": "drg_kategorie.csv",
    "baze":      "drg_baze.csv",
    "skupiny":   "drg_skupiny.csv",
}

def num(v):
    """CSV numeric -> int/float/None (handles '', 'N/A', decimal points)."""
    if v is None:
        return None
    s = str(v).strip().replace("\xa0", "").replace(" ", "")
    if s in ("", "-", "N/A", "n/a", "null"):
        return None
    s = s.replace(",", ".")
    try:
        f = float(s)
    except ValueError:
        return None
    return int(f) if f.is_integer() else round(f, 2)

def write_json(name, data):
    path = OUT / name
    with io.open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    print(f"  wrote {name} ({path.stat().st_size/1024:.1f} KB)")

# ── 1) national indicators ────────────────────────────────────────────────────
print("Part 1: national CZ-DRG indicators", flush=True)
level_counts = {}
for level, fname in LEVELS.items():
    src = SRC / fname
    with io.open(src, encoding="utf-8-sig", errors="replace") as f:
        reader = csv.DictReader(f)
        rows = []
        for row in reader:
            kod = (row.get("drg_kod") or "").strip()
            if not kod:
                continue           # trailing blank lines
            rec = {"kod": kod, "nazev": (row.get("drg_nazev") or "").strip()}
            for key, val in row.items():
                if key in ("drg_kod", "drg_nazev") or key is None:
                    continue
                n = num(val)
                if n is not None:
                    rec[key] = n
            rows.append(rec)
    rows.sort(key=lambda r: -(r.get("pocet_hp_cr") or 0))
    level_counts[level] = len(rows)
    write_json(f"indicators_{level}.json", rows)
    print(f"    {level}: {len(rows)} units")

# ── 2) per-provider case counts ───────────────────────────────────────────────
print("Part 2: per-provider case counts (IČZ x DRG x rok)", flush=True)
providers = {}                    # icz -> name (latest year wins)
units = {}                        # kod -> {kod, nazev, level}
counts = {}                       # kod -> icz -> [y2021..y2024]

for yi, year in enumerate(PROVIDER_YEARS):
    path = SRC / f"drg_icz_{year}.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # locate header block: the row whose first cell names the taxonomy column
    hdr_i = next(i for i, r in enumerate(rows)
                 if r and r[0] and "Taxonomická jednotka" in str(r[0]))
    icz_row, name_row = rows[hdr_i + 1], rows[hdr_i + 2]
    icz_list, name_list = [], []
    for c in range(3, len(icz_row)):
        icz = str(icz_row[c]).strip() if icz_row[c] else ""
        if not icz:
            continue
        icz_list.append((c, icz))
        name_list.append(str(name_row[c]).strip() if c < len(name_row) and name_row[c] else icz)
    for (c, icz), nm in zip(icz_list, name_list):
        providers[icz] = nm       # later years overwrite -> newest name

    n_cells = 0
    for r in rows[hdr_i + 3:]:
        if not r or not r[0] or not r[1]:
            continue
        level = str(r[0]).strip()
        if level not in PROVIDER_LEVELS:
            continue
        kod = str(r[1]).strip()
        units.setdefault(kod, {"kod": kod, "nazev": str(r[2]).strip() if r[2] else kod,
                               "level": "mdc" if level == "MDC" else "baze"})
        for c, icz in icz_list:
            if c >= len(r):
                continue
            v = num(r[c])
            if not v:
                continue
            counts.setdefault(kod, {}).setdefault(icz, [0, 0, 0, 0])[yi] = v
            n_cells += 1
    print(f"    {year}: {len(icz_list)} poskytovatelu, {n_cells:,} nenulovych bunek", flush=True)

provider_list = [{"icz": i, "name": n} for i, n in sorted(providers.items(), key=lambda kv: kv[1])]
unit_list = sorted(units.values(), key=lambda u: (u["level"] != "mdc", u["kod"]))

write_json("providers.json", {
    "years": PROVIDER_YEARS,
    "providers": provider_list,
    "units": unit_list,
    "counts": counts,
})
print(f"    total: {len(provider_list)} poskytovatelu, {len(unit_list)} DRG jednotek, "
      f"{sum(len(v) for v in counts.values()):,} dvojic")

# ── 3) meta ───────────────────────────────────────────────────────────────────
write_json("meta.json", {
    "indicatorYear": INDICATOR_YEAR,
    "providerYears": PROVIDER_YEARS,
    "levelCounts": level_counts,
    "providerCount": len(provider_list),
    "sources": [
        {"id": "1766", "title": "CZ-DRG: Přehled vybraných ukazatelů akutní lůžkové péče",
         "url": "https://www.nzip.cz/data/1766-ukazatele-akutni-luzkova-pece-cz-drg-kategorie-otevrena-data",
         "license": "CC BY 4.0", "year": INDICATOR_YEAR},
        {"id": "2081", "title": "Hospitalizační případy dle IČZ a CZ-DRG",
         "url": "https://www.nzip.cz/data/2081-hospitalizacni-pripady-icz-cz-drg-datovy-souhrn",
         "license": "CC BY 4.0", "year": "2021–2024"},
    ],
})
print("DONE")
